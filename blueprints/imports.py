import os, datetime
from flask import (Blueprint, render_template, request,
                   redirect, url_for, flash, session)
from database import get_db
from blueprints.auth import manager_required, log_action

imports_bp = Blueprint("imports", __name__, url_prefix="/manager/import")

UPLOAD_FOLDER = "/tmp/ra_uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED = {"xlsx", "xls", "csv"}

def allowed(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED


@imports_bp.route("/")
@manager_required
def index():
    db = get_db()
    # Last import timestamps
    last_wo = db.execute(
        "SELECT MAX(created_at) last, COUNT(*) n FROM jobs WHERE source='import'"
    ).fetchone()
    last_so = db.execute(
        "SELECT MAX(created_at) last, COUNT(*) n FROM jobs WHERE sale_value IS NOT NULL AND source='import'"
    ).fetchone()
    recent = db.execute(
        "SELECT job_number, status, created_at FROM jobs ORDER BY created_at DESC LIMIT 10"
    ).fetchall()
    db.close()
    return render_template("manager/imports/index.html",
        last_wo=last_wo, last_so=last_so, recent=recent,
        active_page="imports")


@imports_bp.route("/works-orders", methods=["POST"])
@manager_required
def upload_works_orders():
    f = request.files.get("file")
    if not f or not f.filename or not allowed(f.filename):
        flash("Please upload a valid .xlsx file.", "error")
        return redirect(url_for("imports.index"))

    path = os.path.join(UPLOAD_FOLDER, "works_orders.xlsx")
    f.save(path)

    try:
        count = _import_works_orders(path)
        log_action(session["user_id"], "import_works_orders", f"{count} jobs processed")
        flash(f"Works Orders imported: {count} jobs processed.", "success")
        # Mark readiness check
        db = get_db()
        today = datetime.date.today()
        days_to_mon = (7 - today.weekday()) % 7 or 7
        next_monday = (today + datetime.timedelta(days=days_to_mon)).isoformat()
        db.execute("""UPDATE planning_checks SET checked_at=datetime('now'), resets_on=?
                      WHERE check_name='works_orders'""", (next_monday,))
        db.commit(); db.close()
    except Exception as e:
        flash(f"Import error: {e}", "error")

    return redirect(url_for("imports.index"))


@imports_bp.route("/sales-orders", methods=["POST"])
@manager_required
def upload_sales_orders():
    f = request.files.get("file")
    if not f or not f.filename or not allowed(f.filename):
        flash("Please upload a valid .xlsx file.", "error")
        return redirect(url_for("imports.index"))

    path = os.path.join(UPLOAD_FOLDER, "sales_orders.xlsx")
    f.save(path)

    try:
        count = _import_sales_orders(path)
        log_action(session["user_id"], "import_sales_orders", f"{count} SOs processed")
        flash(f"Sales Orders imported: {count} lines processed.", "success")
        db = get_db()
        today = datetime.date.today()
        days_to_mon = (7 - today.weekday()) % 7 or 7
        next_monday = (today + datetime.timedelta(days=days_to_mon)).isoformat()
        db.execute("""UPDATE planning_checks SET checked_at=datetime('now'), resets_on=?
                      WHERE check_name='sales_orders'""", (next_monday,))
        db.commit(); db.close()
    except Exception as e:
        flash(f"Import error: {e}", "error")

    return redirect(url_for("imports.index"))


# ── Import logic (extracted from original import script) ────────

BOM_PARENTS = {
    'WMS00000027': ['WMU00000016','WMU00000017','WMS00000045','WMS00000049'],
    'WMS00000029': ['WMU00000016','WMU00000017','WMS00000045','WMS00000049'],
    'WMS00000054': ['WMS00000049','WMU00000016','WMU00000017'],
    'WMS00000057': ['WMS00000072','WMS00000073','WMS00000074','WMS00000077',
                    'WMS00000078','WMS00000079','WMS00000080',
                    'WMU00000016','WMU00000017','WMU00000020'],
    'WMS00000022': [],
}
ASSEMBLY_PNS = set(BOM_PARENTS.keys())
IN_PROGRESS_STAGES = {
    '246 - Mechanical Assembly/Wiring',
    '246 - Cut Strip Crimp Print Wires',
    '335 - Dispatch', '330 - Inspection', '220 - Kit Check',
}


def _load_so_map(db):
    """Return part_number -> {earliest_due, total_qty, value, so_ref, customer}."""
    rows = db.execute("""
        SELECT erp_ref so_ref, p.part_number,
               MIN(due_date) earliest_due,
               SUM(planned_qty) total_qty,
               SUM(sale_value)  value,
               customer
        FROM jobs j JOIN parts p ON p.id=j.part_id
        WHERE j.sale_value IS NOT NULL AND j.erp_ref IS NOT NULL
        GROUP BY p.part_number
    """).fetchall()
    return {r["part_number"]: dict(r) for r in rows}


def _sub_due_map(so_map):
    """Return child_pn -> earliest due date inherited from parent console SOs."""
    sub_due = {}
    for parent_pn, children in BOM_PARENTS.items():
        if parent_pn not in so_map: continue
        due = so_map[parent_pn]["earliest_due"]
        if not due: continue
        for child_pn in children:
            if child_pn not in sub_due or due < sub_due[child_pn]:
                sub_due[child_pn] = due
    return sub_due


def _import_sales_orders(path):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    h  = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]

    db      = get_db()
    tier_ids= {r["name"]:r["id"] for r in db.execute("SELECT id,name FROM tiers").fetchall()}
    count   = 0

    for r in range(2, ws.max_row+1):
        row = {h[c-1]: ws.cell(r,c).value for c in range(1, ws.max_column+1)}
        if not row.get("ItemCode") or row.get("Cancelled") == "True": continue

        pn   = str(row["ItemCode"]).strip()
        so   = str(row["Sales Order"]).strip()
        due  = row["Date Required"]
        if isinstance(due, datetime.datetime): due = due.date().isoformat()
        qty  = int(row.get("Qty Outstanding") or 0)
        val  = float(row.get("Sub Total Base Currency") or 0)
        cust = str(row.get("Customer","") or "").strip()
        desc = str(row.get("Item Description","") or "")[:200]

        # Ensure part exists
        pt = "assembly" if pn in ASSEMBLY_PNS else "single"
        db.execute("INSERT OR IGNORE INTO parts(part_number,description,part_type,min_grade_id,estimated_hours) VALUES(?,?,?,?,1.0)",
                   (pn, desc, pt, tier_ids["Competent"]))
        part = db.execute("SELECT id FROM parts WHERE part_number=?", (pn,)).fetchone()
        if not part: continue

        # Upsert job representing the SO line
        existing = db.execute(
            "SELECT id FROM jobs WHERE job_number=? AND erp_ref=?", (so+".SO", so)
        ).fetchone()
        if existing:
            db.execute("""UPDATE jobs SET due_date=?,sale_value=?,planned_qty=?,
                          updated_at=datetime('now') WHERE id=?""",
                       (str(due)[:10] if due else None, val, qty, existing["id"]))
        else:
            db.execute("""INSERT INTO jobs(job_number,part_id,description,customer,quantity,
                                           status,due_date,sale_value,planned_qty,for_stock,team,
                                           erp_ref,source,created_at,updated_at)
                          VALUES(?,?,?,?,?,'Unscheduled',?,?,?,0,'General',?,
                                 'sales_order',datetime('now'),datetime('now'))""",
                       (so+".SO", part["id"], desc, cust, qty,
                        str(due)[:10] if due else None, val, qty, so))
        count += 1

    # Update WO due dates from new SO data
    so_map  = _load_so_map(db)
    sub_due = _sub_due_map(so_map)
    for pn, info in so_map.items():
        db.execute("""UPDATE jobs SET due_date=?, sale_value=?
                      WHERE part_id=(SELECT id FROM parts WHERE part_number=?)
                        AND source='import' AND due_date IS NULL""",
                   (info["earliest_due"], info["value"], pn))
    for pn, due in sub_due.items():
        db.execute("""UPDATE jobs SET due_date=?
                      WHERE part_id=(SELECT id FROM parts WHERE part_number=?)
                        AND source='import' AND due_date IS NULL""",
                   (due, pn))
    # Propagate planned_qty (SO urgent qty) to matching WOs by part number
    for pn, info in so_map.items():
        total_so_qty = info.get('total_qty') or 0
        if total_so_qty:
            db.execute("""
                UPDATE jobs SET planned_qty=?
                WHERE source='import'
                  AND part_id=(SELECT id FROM parts WHERE part_number=?)
                  AND (planned_qty IS NULL OR planned_qty != ?)
            """, (total_so_qty, pn, total_so_qty))
    db.commit(); db.close()
    return count


def _import_works_orders(path):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    h  = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]

    db       = get_db()
    tier_ids = {r["name"]:r["id"] for r in db.execute("SELECT id,name FROM tiers").fetchall()}
    so_map   = _load_so_map(db)
    sub_due  = _sub_due_map(so_map)
    inserted = updated = 0

    for r in range(2, ws.max_row+1):
        row = {h[c-1]: ws.cell(r,c).value for c in range(1, ws.max_column+1)}
        if row.get("Cancelled") == "True": continue
        if not row.get("Outstanding") or int(row.get("Outstanding") or 0) <= 0: continue

        wo_num   = str(row["Transaction No"]).strip()
        pn       = str(row.get("Item","") or "").strip()
        desc     = str(row.get("Description1","") or "")[:200]
        qty      = int(row["Outstanding"])
        project  = str(row.get("Project","") or "").strip()
        customer = str(row.get("CustomerSupplier","") or "").strip()
        stage    = str(row.get("Current Task","") or "").strip()
        so_ref   = str(row.get("Sales Order","") or "").strip() or None

        team      = "Davenham" if "Davenham" in customer else "General"
        for_stock = 1 if project in ("Stock Build","") or customer in ("Rapid Assembly Stock","") else 0
        status    = "In Progress" if stage in IN_PROGRESS_STAGES else "Unscheduled"
        pt        = "assembly" if pn in ASSEMBLY_PNS else "single"
        eas       = project if project and project not in ("Stock Build","") else None

        # Due date from SO cross-reference
        so_info = so_map.get(pn)
        if so_info:
            due_date   = so_info["earliest_due"]
            sale_value = so_info["value"]
            so_qty     = so_info["total_qty"]
            so_num     = so_info["so_ref"]
        elif pn in sub_due:
            due_date   = sub_due[pn]
            sale_value = None; so_qty = None; so_num = so_ref
        else:
            wo_due = row.get("Delivery Date")
            if isinstance(wo_due, datetime.datetime): wo_due = wo_due.date().isoformat()
            due_date   = str(wo_due)[:10] if wo_due else None
            sale_value = None; so_qty = None; so_num = so_ref

        db.execute("INSERT OR IGNORE INTO parts(part_number,description,part_type,min_grade_id,estimated_hours) VALUES(?,?,?,?,1.0)",
                   (pn, desc, pt, tier_ids["Competent"]))
        part = db.execute("SELECT id FROM parts WHERE part_number=?", (pn,)).fetchone()
        if not part: continue

        existing = db.execute("SELECT id,status FROM jobs WHERE job_number=?", (wo_num,)).fetchone()
        if existing:
            # Update qty (outstanding may have reduced since last import),
            # due date and status — but don't override Complete
            if existing["status"] != "Complete":
                db.execute("""UPDATE jobs SET quantity=?,status=?,due_date=?,
                              sale_value=?,updated_at=datetime('now') WHERE id=?""",
                           (qty, status, due_date, sale_value, existing["id"]))
                updated += 1
            continue

        db.execute("""INSERT INTO jobs(job_number,part_id,description,customer,quantity,
                                       status,due_date,sale_value,planned_qty,for_stock,team,
                                       eas_number,erp_ref,source,created_at,updated_at)
                      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,'import',datetime('now'),datetime('now'))""",
                   (wo_num, part["id"], desc, customer[:100], qty,
                    status, due_date, sale_value, so_qty, for_stock, team, eas, so_num))
        if so_qty:
            jid = db.execute("SELECT id FROM jobs WHERE job_number=?", (wo_num,)).fetchone()["id"]
            db.execute("UPDATE jobs SET planned_qty=? WHERE id=?", (so_qty, jid))
        inserted += 1

    db.commit(); db.close()
    return inserted + updated
